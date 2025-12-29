#!/usr/bin/env python3
# ruff: noqa: S101
"""
history.py のテスト
"""
import datetime
import pathlib
import unittest.mock

import openpyxl
import pytest

import merhist.config
import merhist.handle
import merhist.history
import merhist.item


class TestSheetDef:
    """SHEET_DEF のテスト"""

    def test_bought_sheet_def_exists(self):
        """購入シート定義が存在"""
        assert "BOUGHT" in merhist.history.SHEET_DEF
        assert "SHEET_TITLE" in merhist.history.SHEET_DEF["BOUGHT"]
        assert "TABLE_HEADER" in merhist.history.SHEET_DEF["BOUGHT"]

    def test_sold_sheet_def_exists(self):
        """販売シート定義が存在"""
        assert "SOLD" in merhist.history.SHEET_DEF
        assert "SHEET_TITLE" in merhist.history.SHEET_DEF["SOLD"]
        assert "TABLE_HEADER" in merhist.history.SHEET_DEF["SOLD"]

    def test_bought_columns(self):
        """購入シートのカラム定義"""
        cols = merhist.history.SHEET_DEF["BOUGHT"]["TABLE_HEADER"]["col"]
        expected_cols = ["shop_name", "date", "name", "image", "count", "price", "condition"]
        for col in expected_cols:
            assert col in cols

    def test_sold_columns(self):
        """販売シートのカラム定義"""
        cols = merhist.history.SHEET_DEF["SOLD"]["TABLE_HEADER"]["col"]
        expected_cols = ["shop_name", "date", "name", "price", "commission", "postage", "profit"]
        for col in expected_cols:
            assert col in cols

    def test_link_func_bought(self):
        """購入シートのリンク関数"""
        cols = merhist.history.SHEET_DEF["BOUGHT"]["TABLE_HEADER"]["col"]
        item = {"url": "https://example.com/item", "order_url": "https://example.com/order"}

        assert cols["id"]["link_func"](item) == "https://example.com/item"
        assert cols["no"]["link_func"](item) == "https://example.com/order"

    def test_link_func_sold(self):
        """販売シートのリンク関数"""
        cols = merhist.history.SHEET_DEF["SOLD"]["TABLE_HEADER"]["col"]
        item = {"url": "https://example.com/item", "id": "m123", "shop": "mercari.com"}

        assert cols["id"]["link_func"](item) == "https://example.com/item"

    def test_conv_func_commission_rate(self):
        """手数料率の変換関数"""
        cols = merhist.history.SHEET_DEF["SOLD"]["TABLE_HEADER"]["col"]
        assert cols["commission_rate"]["conv_func"](10) == 0.1
        assert cols["commission_rate"]["conv_func"](5) == 0.05


class TestGenerateSheet:
    """generate_sheet のテスト"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """モック Config"""
        config = unittest.mock.MagicMock(spec=merhist.config.Config)
        config.cache_file_path = tmp_path / "cache" / "order.pickle"
        config.selenium_data_dir_path = tmp_path / "selenium"
        config.debug_dir_path = tmp_path / "debug"
        config.thumb_dir_path = tmp_path / "thumb"
        config.captcha_file_path = tmp_path / "captcha.png"
        config.excel_file_path = tmp_path / "output" / "mercari.xlsx"
        config.excel_font = openpyxl.styles.Font(name="Arial", size=10)
        return config

    @pytest.fixture
    def handle(self, mock_config):
        """Handle インスタンス"""
        with unittest.mock.patch("my_lib.serializer.load", return_value=merhist.handle.TradingInfo()):
            h = merhist.handle.Handle(config=mock_config)
            h.progress_manager = unittest.mock.MagicMock()
            h.progress_bar = {}
            return h

    def test_generate_sheet_empty(self, handle):
        """空のリストでシート生成"""
        book = openpyxl.Workbook()

        with unittest.mock.patch("my_lib.openpyxl_util.generate_list_sheet") as mock_gen:
            merhist.history.generate_sheet(handle, book, is_need_thumb=True)

            assert mock_gen.call_count == 2

    def test_generate_sheet_with_items(self, handle):
        """アイテムありでシート生成"""
        book = openpyxl.Workbook()

        bought_item = merhist.item.BoughtItem(id="m123", name="購入商品")
        sold_item = merhist.item.SoldItem(id="s456", name="販売商品", price=1000)

        handle.trading.bought_item_list = [bought_item]
        handle.trading.sold_item_list = [sold_item]

        with unittest.mock.patch("my_lib.openpyxl_util.generate_list_sheet") as mock_gen:
            merhist.history.generate_sheet(handle, book, is_need_thumb=True)

            assert mock_gen.call_count == 2

    def test_generate_sheet_without_thumb(self, handle):
        """サムネイルなしでシート生成"""
        book = openpyxl.Workbook()

        with unittest.mock.patch("my_lib.openpyxl_util.generate_list_sheet") as mock_gen:
            merhist.history.generate_sheet(handle, book, is_need_thumb=False)

            # is_need_thumb=False で呼ばれることを確認
            for call in mock_gen.call_args_list:
                assert call[0][3] is False  # is_need_thumb


class TestGenerateTableExcel:
    """generate_table_excel のテスト"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """モック Config"""
        config = unittest.mock.MagicMock(spec=merhist.config.Config)
        config.cache_file_path = tmp_path / "cache" / "order.pickle"
        config.selenium_data_dir_path = tmp_path / "selenium"
        config.debug_dir_path = tmp_path / "debug"
        config.thumb_dir_path = tmp_path / "thumb"
        config.captcha_file_path = tmp_path / "captcha.png"
        config.excel_file_path = tmp_path / "output" / "mercari.xlsx"
        config.excel_font = openpyxl.styles.Font(name="Arial", size=10)
        return config

    @pytest.fixture
    def handle(self, mock_config):
        """Handle インスタンス"""
        with unittest.mock.patch("my_lib.serializer.load", return_value=merhist.handle.TradingInfo()):
            h = merhist.handle.Handle(config=mock_config)
            h.progress_manager = unittest.mock.MagicMock()
            mock_counter = unittest.mock.MagicMock()
            h.progress_bar = {
                merhist.history.STATUS_ALL: mock_counter,
                merhist.history.STATUS_INSERT_ITEM: mock_counter,
            }
            return h

    def test_generate_table_excel(self, handle, tmp_path):
        """Excelファイル生成（ダミーデータ使用）"""
        excel_path = tmp_path / "output" / "test.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        # ダミーデータを追加
        bought_item = merhist.item.BoughtItem(
            id="m123", name="テスト購入商品", price=1000,
            purchase_date=datetime.datetime(2025, 1, 15)
        )
        sold_item = merhist.item.SoldItem(
            id="s456", name="テスト販売商品", price=2000,
            completion_date=datetime.datetime(2025, 1, 20)
        )
        handle.trading.bought_item_list = [bought_item]
        handle.trading.sold_item_list = [sold_item]

        merhist.history.generate_table_excel(handle, excel_path, is_need_thumb=False)

        # ファイルが生成されることを確認
        assert excel_path.exists()

        # Excelファイルを開いてシートを確認
        book = openpyxl.load_workbook(excel_path)
        assert len(book.worksheets) == 2
        book.close()

    def test_generate_table_excel_empty(self, handle, tmp_path):
        """空データでExcelファイル生成"""
        excel_path = tmp_path / "output" / "test_empty.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        merhist.history.generate_table_excel(handle, excel_path, is_need_thumb=False)

        assert excel_path.exists()
        book = openpyxl.load_workbook(excel_path)
        assert len(book.worksheets) == 2
        book.close()

    def test_generate_table_excel_normalizes_data(self, handle, tmp_path):
        """データの正規化が呼ばれる"""
        excel_path = tmp_path / "output" / "test_normalize.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        # 重複データを追加
        item1 = merhist.item.BoughtItem(id="m1", name="商品1")
        item2 = merhist.item.BoughtItem(id="m1", name="商品1更新")  # 同じID
        handle.trading.bought_item_list = [item1, item2]
        handle.trading.bought_checked_count = 2

        merhist.history.generate_table_excel(handle, excel_path, is_need_thumb=False)

        # 正規化により重複が削除される
        assert len(handle.trading.bought_item_list) == 1

    def test_generate_table_excel_updates_progress(self, handle, tmp_path):
        """プログレスバーが更新される"""
        excel_path = tmp_path / "output" / "test_progress.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        merhist.history.generate_table_excel(handle, excel_path, is_need_thumb=False)

        # プログレスバーが更新されることを確認
        assert handle.progress_bar[merhist.history.STATUS_ALL].update.call_count >= 3


class TestWarningHandler:
    """_warning_handler のテスト"""

    def test_warning_with_name_and_date(self, caplog):
        """商品名と日付がある場合の警告メッセージ"""
        import logging

        item = {
            "name": "テスト商品",
            "purchase_date": datetime.datetime(2025, 1, 15),
        }

        with caplog.at_level(logging.WARNING):
            merhist.history._warning_handler(item, "テスト警告")

        assert len(caplog.records) == 1
        assert "⚠️" in caplog.text
        assert "25年01月15日" in caplog.text
        assert "テスト商品" in caplog.text
        assert "テスト警告" in caplog.text

    def test_warning_without_name(self, caplog):
        """商品名がない場合は「不明」を使用"""
        import logging

        item = {
            "purchase_date": datetime.datetime(2025, 1, 15),
        }

        with caplog.at_level(logging.WARNING):
            merhist.history._warning_handler(item, "テスト警告")

        assert "不明" in caplog.text
        assert "テスト警告" in caplog.text

    def test_warning_without_date(self, caplog):
        """日付がない場合は日付部分を省略"""
        import logging

        item = {
            "name": "テスト商品",
        }

        with caplog.at_level(logging.WARNING):
            merhist.history._warning_handler(item, "テスト警告")

        assert "テスト商品: テスト警告" in caplog.text
        assert "年" not in caplog.text.split("テスト商品")[0]  # 日付がないことを確認

    def test_warning_with_none_date(self, caplog):
        """日付が None の場合は日付部分を省略"""
        import logging

        item = {
            "name": "テスト商品",
            "purchase_date": None,
        }

        with caplog.at_level(logging.WARNING):
            merhist.history._warning_handler(item, "テスト警告")

        assert "テスト商品: テスト警告" in caplog.text

    def test_warning_with_non_datetime_date(self, caplog):
        """日付が datetime でない場合は日付部分を省略"""
        import logging

        item = {
            "name": "テスト商品",
            "purchase_date": "2025-01-15",  # 文字列
        }

        with caplog.at_level(logging.WARNING):
            merhist.history._warning_handler(item, "テスト警告")

        assert "テスト商品: テスト警告" in caplog.text
        # 日付形式でないので年月日が含まれない
        assert "年" not in caplog.text.split("テスト商品")[0]


class TestStatusConstants:
    """ステータス定数のテスト"""

    def test_status_insert_item(self):
        """STATUS_INSERT_ITEM が定義されている"""
        assert merhist.history.STATUS_INSERT_ITEM
        assert "Insert" in merhist.history.STATUS_INSERT_ITEM

    def test_status_all(self):
        """STATUS_ALL が定義されている"""
        assert merhist.history.STATUS_ALL
        assert "Excel" in merhist.history.STATUS_ALL

    def test_shop_name(self):
        """SHOP_NAME が定義されている"""
        assert merhist.history.SHOP_NAME == "メルカリ"
